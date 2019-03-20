from openpyxl import load_workbook
from  common.peizhi_file import ReadConfig
class Cases:  #第三种方法加的类
    '''专门存储测试数据'''
    def __init__(self):
        self.case_id = None
        self.title = None
        self.a = None
        self.b = None
        self.expected = None

class DoExcel:
    '''来获取excel对应表单数据的类 '''
    def  __init__(self, file_name, sheet_name):
        self.file_name = file_name  # 工作簿
        self.sheet_name = sheet_name  # 表单名
        self.wb = load_workbook(self.file_name)
        self.sheet = self.wb[self.sheet_name]  #这两行在下面的方法中都用到了，所以把他放到初始化函数里面会比较方便
        self.button = ReadConfig('conf/math_method.conf').get_value('Case', 'button')

    def  get_data(self):  # 获取数据
        # #方法一
        # # 读取数据的具体操作
        # test_data = []  # 所有的数据都存在
        # for i in range(2, self.sheet.max_row + 1):  # 以行开始读，从第二行开始
        #     row_data = []  # 每一行数据存到这个列表里面
        #     row_data.append(self.sheet.cell(i, 1).value)  # 存case_id
        #     row_data.append(self.sheet.cell(i, 2).value)  # 存title
        #     row_data.append(self.sheet.cell(i, 3).value)  # 存a
        #     row_data.append(self.sheet.cell(i, 4).value)  # 存b
        #     row_data.append(self.sheet.cell(i, 5).value)  # 存expected
        #     test_data.append(row_data)
        #
        # # 使用配置文件来决定执行哪些用例
        # if self.button == 'all':
        #     return test_data    # 返回的[ [ ],[ ],[ ] ]
        # else:
        #     list = eval(self.button)
        #     final_data =[]
        #     for i in list:
        #         final_data.append(test_data[i-1])
        #     return final_data

        # #方法二
        # # 读取数据的具体操作
        # test_data = []  # 所有的数据都存在
        # for i in range(2, self.sheet.max_row+1):  # 以行开始读，从第二行开始
        #     row_data = {}   # 每一行数据存到这个字典里面
        #     row_data['case_id'] = self.sheet.cell(i,1).value  #存case_id
        #     row_data['title'] = self.sheet.cell(i,2).value  #存title
        #     row_data['a'] = self.sheet.cell(i, 3).value   #存a
        #     row_data['b'] = self.sheet.cell(i, 4).value   #存b
        #     row_data['expected'] = self.sheet.cell(i, 5).value  #存expected
        #     test_data.append(row_data)
        #
        # # 使用配置文件来决定执行哪些用例
        # if self.button == 'all':
        #     return test_data  # 返回的[{  },{ },{ }]
        # else:
        #     list = eval(self.button)
        #     final_data = []
        #     for i in list:
        #         final_data.append(test_data[i - 1])
        #     return final_data


        # 方法三  （重要） 类与对象的思想
        # 读取数据的具体操作
        case = []  # 存储数据
        for i in range(2, self.sheet.max_row + 1):  # 以行开始读，从第二行开始
            row_case = Cases()  # 每一行数据存到这个对象里面   对象/实例
            row_case.case_id = self.sheet.cell(i, 1).value  # 存case_id
            row_case.title = self.sheet.cell(i, 2).value  # 存title
            row_case.a = self.sheet.cell(i, 3).value  # 存a
            row_case.b = self.sheet.cell(i, 4).value  # 存b
            row_case.expected = self.sheet.cell(i, 5).value  # 存expected
            case.append(row_case)

        # 使用配置文件来决定执行哪些用例
        if self.button == 'all':
            return case   #返回的是[ 对象值，对象值，对象值，对象值]  每一行的所有元素是一个对象的属性，
        else:
            list = eval(self.button)
            final_case =[]
            for i in list:
                final_case.append(case[i-1])
            return final_case


    def write_back(self, row, column,value):  # 写回数据  行 列 值

        res = self.sheet.cell(row,column).value = value  # 写入值

        self.wb.save(self.file_name)  # 保存结果（关闭excel的情况下）


if __name__ == '__main__':
    test_data = DoExcel('method_message.xlsx', 'add',).get_data()
    print(test_data)