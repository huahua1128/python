from openpyxl import load_workbook   #读写(load_workbook 是函数）
from openpyxl import Workbook   #新建(Workbook 是类)

# 新建一个excel  (.xlsx)
# wb=Workbook()  #新建一个excel
# wb.save("course_1229.xlsx")   #保存

# load_workbook 函数   来进行excel 文件的读写
# 1. 打开excel
wb=load_workbook("course_1229.xlsx")

# 2. 定位表单
# sheet=wb.get_sheet_by_name('test_name') # 方法一  已经过时，一般不用
sheet=wb['test_data']   #方法二  现在用的

# 3.定位单元格---取值
res=sheet.cell(row = 3, column = 4).value  # row行  column列
# print(res)

# 3.定位单元格---写值
#方法一
sheet.cell(row = 3, column = 5, value = 999)
wb.save("course_1229.xlsx")   #写值要保存
#涉及到写的时候一定要把excel关掉。
#方法二
sheet.cell(row = 4, column = 6).value='777777'
wb.save("course_1229.xlsx")

# res = sheet.cell(row = 4, column = 2).value
# print("获取到的值是：{}".format(res))
# print("获取到的数据类型是：{}".format(type(res)))
# res = eval(sheet.cell(row = 4, column = 2).value)
# print("获取到的值是：{}".format(res))
# print("获取到的数据类型是：{}".format(type(res)))

#读取excel中的所有值
# print("行", sheet.max_row)  #获取最大行（有数据的）
# print("列", sheet.max_column)  #获取最大列（有数据的）

#读取每个单元格的值
for  i  in  range(1,sheet.max_row +1):
    for  j  in  range(1,sheet.max_column +1):
        res=sheet.cell(row = i, column = j).value
        print("获取到的值是：{}".format(res))