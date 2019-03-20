from  openpyxl import  load_workbook
class  Excel_Operation:

    def  __init__(self,excel_name,sheet_name):
        self.excel_name = excel_name
        self.sheet_name = sheet_name

    def  read_data(self):
        wb = load_workbook(self.excel_name)
        sheet = wb[self.sheet_name]
        list = []
        for i  in  range(1, sheet.max_column + 1):
            list.append(sheet.cell(row = 1, column = i).value)

        list_data = []
        for i  in  range(2,sheet.max_row + 1):
            dict = {}
            for j  in range(1,4):
                dict[list[j-1]]=sheet.cell(row=i, column=j).value
            list_data.append(dict)

        return list_data

    def write_data(self,row,column,value):
        wb = load_workbook(self.excel_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row = row, column = column).value = value
        wb.save(self.excel_name)

if  __name__=='__main__':
    res=Excel_Operation('method_message.xlsx','add').read_data()
    print(res)